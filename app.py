import re
import numpy as np
import pandas as pd
import streamlit as st
from io import BytesIO
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Validador de Duplicados", layout="centered")
st.title("🎩✨ Validador de Duplicados")
st.write("Suba uma planilha Excel ou informe o link público do Google Sheets para validar duplicados.")

# ---------------- Funções ----------------
def ler_planilha(caminho_ou_link):
    if not isinstance(caminho_ou_link, str):
        return pd.read_excel(caminho_ou_link)

    if caminho_ou_link.startswith("http"):
        if "docs.google.com/spreadsheets" in caminho_ou_link:
            try:
                sheet_id = caminho_ou_link.split("/d/")[1].split("/")[0]
                export_link = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
                resp = requests.get(export_link)
                if resp.status_code == 200:
                    return pd.read_excel(BytesIO(resp.content))
                else:
                    st.error(f"❌ Erro ao acessar o link. Status: {resp.status_code}")
                    return None
            except Exception as e:
                st.error(f"❌ Não foi possível processar o link: {e}")
                return None
        else:
            st.error("❌ O link não parece ser do Google Sheets.")
            return None
    else:
        return pd.read_excel(caminho_ou_link)


def _detectar_coluna(df, termos):
    """Detecta primeira coluna cujo nome contenha qualquer um dos termos (case-insensitive)."""
    lc = [c for c in df.columns]
    for t in termos:
        for c in lc:
            if t in c.lower():
                return c
    return None


def _parse_valor(v):
    """Tenta converter diferentes formatos de valor para float (R$, 9.600,00, 9600.00, etc.)."""
    if pd.isna(v):
        return None
    s = str(v).strip()
    # remover símbolos de moeda e espaços
    s = s.replace("R$", "").replace("r$", "").replace(" ", "")
    # heurística brasileira: se contém vírgula, tratar ',' como separador decimal
    if "," in s:
        # remover pontos (milhares) e trocar vírgula por ponto
        s2 = s.replace(".", "").replace(",", ".")
    else:
        s2 = s.replace(",", "")
    # retirar caracteres que não sejam dígito, ponto ou sinal
    s2 = re.sub(r"[^\d\.\-+]", "", s2)
    try:
        return float(s2)
    except:
        return None


def _normalize_cliente(v):
    if pd.isna(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(int(v))
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip().lower()


def _normalizar_data_valor_cliente(row, date_col, client_col, value_col):
    # data: converter para datetime e pegar .date() (ignorando hora)
    raw_date = row.get(date_col, None)
    try:
        dt = pd.to_datetime(raw_date, dayfirst=True, errors="coerce")
        d = dt.date() if not pd.isna(dt) else None
    except:
        d = None

    cliente = _normalize_cliente(row.get(client_col, None))
    valor = _parse_valor(row.get(value_col, None))
    return d, cliente, valor


def marcar_duplicados_vermelho(df):
    # Detectar colunas mais prováveis
    date_col = _detectar_coluna(df, ["data", "carimbo", "timestamp", "date"]) or df.columns[0]
    client_col = _detectar_coluna(df, ["cliente", "client", "cod", "codigo"]) or None
    value_col = _detectar_coluna(df, ["valor", "value", "amount", "total"]) or None

    # Se não detectou cliente ou valor, tenta escolher colunas razoáveis
    if client_col is None:
        # procurar coluna com dtype int/str que pareça ser código
        for c in df.columns:
            if "empresa" in c.lower():  # evitar pegar empresa por engano
                continue
            if df[c].dtype == object or pd.api.types.is_integer_dtype(df[c]) or "cod" in c.lower():
                client_col = c
                break
        if client_col is None:
            # fallback: coluna 2
            client_col = df.columns[2] if len(df.columns) > 2 else df.columns[0]

    if value_col is None:
        # procurar primeira coluna numérica ou com "valor" no nome
        for c in df.columns:
            if pd.api.types.is_numeric_dtype(df[c]):
                value_col = c
                break
        if value_col is None:
            value_col = df.columns[3] if len(df.columns) > 3 else df.columns[-1]

    st.write(f"Usando colunas para verificar duplicados: Data = **{date_col}**, Cliente = **{client_col}**, Valor = **{value_col}**")

    # Garantir coluna de sinalização
    df = df.copy()
    df["Duplicado_Linha"] = ""

    primeira_ocorrencia = {}

    # Preencher coluna Duplicado_Linha (ignorando horário — usamos apenas .date())
    for idx, row in df.iterrows():
        d, cliente_norm, valor_num = _normalizar_data_valor_cliente(row, date_col, client_col, value_col)
        key = (d, cliente_norm, None if valor_num is None else round(valor_num, 2))
        if key in primeira_ocorrencia and key[0] is not None and key[2] is not None and key[1] != "":
            first_idx = primeira_ocorrencia[key]
            df.at[idx, "Duplicado_Linha"] = f"Primeira ocorrência na linha {first_idx + 2}"
        else:
            primeira_ocorrencia[key] = idx

    # Salvar temporário em Excel e usar openpyxl para pintar
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    col_dup = df.columns.get_loc("Duplicado_Linha") + 1

    # Pintar apenas linhas que têm comentário na coluna Duplicado_Linha
    for row_idx in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row_idx, column=col_dup).value
        if cell_value and str(cell_value).strip() != "":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col).fill = vermelho

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    qtd_dup = (df["Duplicado_Linha"] != "").sum()
    return final_output, qtd_dup


# ---------------- Interface ----------------
tab1, tab2 = st.tabs(["📂 Upload Excel", "🔗 Link Google Sheets"])
df = None

with tab1:
    uploaded_file = st.file_uploader("Selecione um arquivo Excel", type=["xlsx"])
    if uploaded_file is not None:
        df = ler_planilha(uploaded_file)

with tab2:
    link = st.text_input("Cole o link público do Google Sheets:")
    if link:
        df = ler_planilha(link)

if df is not None:
    st.subheader("📑 Pré-visualização dos dados")
    st.dataframe(df.head())

    if st.button("🔎 Validar Duplicados"):
        arquivo_final, qtd_dup = marcar_duplicados_vermelho(df)

        if qtd_dup > 0:
            st.success(f"✅ Foram encontradas {qtd_dup} linhas duplicadas (segunda ocorrência em diante).")
        else:
            st.info("Nenhuma linha duplicada encontrada.")

        st.download_button(
            label="📥 Baixar planilha validada",
            data=arquivo_final,
            file_name="planilha_validada.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
